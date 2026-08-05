import asyncio
import csv
import io
import json
from typing import Literal

from fastapi import APIRouter, Depends, File, Form, HTTPException, Query, UploadFile, status
from fastapi.responses import StreamingResponse

from app.analysis_repository import AnalysisRepository
from app.analysis_service import FileParsingError, FileValidationError, clean_and_analyze, normalize_cleaning_options, parse_file, validate_upload
from app.report_service import build_pdf_report
from app.user_store import User


def build_analysis_router(repository: AnalysisRepository, current_user) -> APIRouter:
    router = APIRouter(prefix='/api/v1', tags=['data quality'])

    async def owned(analysis_id: str, user: User):
        analysis = await repository.find_for_user(analysis_id, user['id'])
        if analysis is None:
            raise HTTPException(status_code=404, detail='Analysis not found.')
        return analysis

    @router.post('/files/analyze', status_code=status.HTTP_201_CREATED)
    async def analyze(file: UploadFile = File(...), cleaning_config: str = Form('{}'), user: User = Depends(current_user)):
        content = await file.read()
        try:
            filename, file_type = validate_upload(file.filename, content)
            try:
                supplied = json.loads(cleaning_config or '{}')
            except json.JSONDecodeError as error:
                raise FileValidationError('Cleaning configuration must be valid JSON.') from error
            if not isinstance(supplied, dict):
                raise FileValidationError('Cleaning configuration must be a JSON object.')
            records = await asyncio.to_thread(parse_file, filename, content)
            cleaned, metrics, charts = await asyncio.to_thread(clean_and_analyze, records, normalize_cleaning_options(supplied))
        except OverflowError as error:
            raise HTTPException(status_code=413, detail=str(error)) from error
        except FileValidationError as error:
            raise HTTPException(status_code=400, detail=str(error)) from error
        except FileParsingError as error:
            raise HTTPException(status_code=422, detail=str(error)) from error
        job = await repository.create(user['id'], filename, file_type, len(content), cleaned, metrics, charts)
        job.pop('records', None)
        return job

    @router.get('/analyses')
    async def analyses(page: int = Query(1, ge=1), page_size: int = Query(20, ge=1, le=100), user: User = Depends(current_user)):
        total, items = await asyncio.gather(repository.count_for_user(user['id']), repository.list_for_user(user['id'], (page - 1) * page_size, page_size))
        return {'page': page, 'page_size': page_size, 'total': total, 'pages': max(1, (total + page_size - 1) // page_size), 'items': items}

    @router.get('/analyses/{analysis_id}')
    async def analysis(analysis_id: str, user: User = Depends(current_user)):
        result = await owned(analysis_id, user)
        result.pop('records', None)
        return result

    @router.get('/analyses/{analysis_id}/records')
    async def records(analysis_id: str, page: int = Query(1, ge=1), page_size: int = Query(50, ge=1, le=500), search: str = Query('', max_length=200), user: User = Depends(current_user)):
        result = await owned(analysis_id, user)
        items = result['records']
        if search.strip():
            needle = search.casefold().strip()
            items = [row for row in items if any(needle in str(value).casefold() for value in row.values())]
        start = (page - 1) * page_size
        return {'page': page, 'page_size': page_size, 'total': len(items), 'columns': result['columns'], 'items': items[start:start + page_size]}

    @router.delete('/analyses/{analysis_id}', status_code=204)
    async def delete(analysis_id: str, user: User = Depends(current_user)):
        if not await repository.delete_for_user(analysis_id, user['id']):
            raise HTTPException(status_code=404, detail='Analysis not found.')

    @router.get('/analyses/{analysis_id}/export/{format}')
    async def export(analysis_id: str, format: Literal['csv', 'xlsx', 'json', 'pdf'], user: User = Depends(current_user)):
        result = await owned(analysis_id, user)
        filename = f"{str(result['filename']).rsplit('.', 1)[0]}_cleaned.{format}"
        if format == 'csv':
            output = io.StringIO(); writer = csv.DictWriter(output, fieldnames=result['columns'], extrasaction='ignore')
            writer.writeheader(); writer.writerows(result['records'])
            return download(output.getvalue().encode('utf-8-sig'), 'text/csv; charset=utf-8', filename)
        if format == 'json':
            payload = {'metadata': {key: result[key] for key in ('analysis_id', 'filename', 'file_type', 'created_at')}, 'metrics': result['metrics'], 'records': result['records']}
            return download(json.dumps(payload, ensure_ascii=False, indent=2).encode(), 'application/json', filename)
        if format == 'pdf':
            return download(await asyncio.to_thread(build_pdf_report, result), 'application/pdf', filename)
        return download(_workbook(result), 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet', filename)

    @router.get('/dashboard/summary')
    async def summary(user: User = Depends(current_user)):
        value = await repository.dashboard_summary(user['id']); total = int(value['cleaned_records'])
        response = {'total_analyses': value['total_analyses'], 'total_records': total, **value}
        for name in ('success', 'failed', 'warning', 'unknown'):
            response[f'{name}_percentage'] = round(int(value[f'{name}_count']) / total * 100, 2) if total else 0.0
        return response

    @router.get('/dashboard/status-distribution')
    async def status_distribution(user: User = Depends(current_user)):
        value = await repository.dashboard_summary(user['id'])
        return {'data': [{'name': name, 'value': value[f'{name.casefold()}_count']} for name in ('Success', 'Failed', 'Warning', 'Unknown')]}

    @router.get('/dashboard/file-type-distribution')
    async def file_types(user: User = Depends(current_user)):
        return {'data': await repository.file_type_distribution(user['id'])}

    @router.get('/dashboard/trends')
    async def trends(range: Literal['7d', '30d', '90d'] = '30d', user: User = Depends(current_user)):
        return {'range': range, 'data': await repository.trends(user['id'], int(range[:-1]))}

    @router.get('/dashboard/recent-analyses')
    async def recent(limit: int = Query(5, ge=1, le=20), user: User = Depends(current_user)):
        items = await repository.list_for_user(user['id'], 0, limit)
        return {'items': [_dashboard_item(item) for item in items]}

    @router.get('/dashboard/latest-analysis')
    async def latest(user: User = Depends(current_user)):
        item = await repository.latest_for_user(user['id'])
        return {'analysis': _dashboard_item(item) if item else None}

    return router


def _dashboard_item(item):
    if not item: return None
    metrics = item['metrics']
    return {'id': item['analysis_id'], 'analysis_id': item['analysis_id'], 'filename': item['filename'],
            'file_type': item['file_type'], 'file_size': item['file_size'], 'status': item['processing_status'],
            'processing_status': item['processing_status'], 'total_records': metrics.get('total_records', 0),
            'quality_score': metrics.get('quality_score', 0), 'duplicates_removed': metrics.get('duplicates_removed', 0),
            'created_at': item['created_at']}


def download(content: bytes, media_type: str, filename: str):
    return StreamingResponse(iter([content]), media_type=media_type,
                             headers={'Content-Disposition': f'attachment; filename="{filename}"', 'Cache-Control': 'no-store'})


def _workbook(analysis) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    workbook = Workbook(); sheet = workbook.active; sheet.title = 'Cleaned_Data'; sheet.append(analysis['columns'])
    for cell in sheet[1]: cell.font = Font(bold=True, color='FFFFFF'); cell.fill = PatternFill('solid', fgColor='12372A')
    for row in analysis['records']: sheet.append([row.get(column, '') for column in analysis['columns']])
    sheet.freeze_panes = 'A2'; sheet.auto_filter.ref = sheet.dimensions
    summary = workbook.create_sheet('Quality_Summary')
    for key, value in analysis['metrics'].items():
        if key not in {'column_profiles', 'warnings', 'cleaning_options'}: summary.append([key, json.dumps(value) if isinstance(value, (dict, list)) else value])
    profiles = workbook.create_sheet('Column_Profile'); profiles.append(['column', 'type', 'missing', 'missing_%', 'unique', 'outliers', 'sample_values'])
    for item in analysis['metrics'].get('column_profiles', []): profiles.append([item.get('name'), item.get('inferred_type'), item.get('missing_count'), item.get('missing_percentage'), item.get('unique_count'), item.get('outlier_count'), ', '.join(item.get('sample_values', []))])
    output = io.BytesIO(); workbook.save(output); return output.getvalue()
